import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import requests
from openpyxl.utils import range_boundaries
import numpy as np
import xlsxwriter

# -------------------------------------------
# Streamlit UI
# -------------------------------------------

st.title("Personal Timetable Creator - Trimester 5, 2nd Half")

st.markdown("""
**Step 1:** Select your campus and subjects from the list  
**Step 2:** Click on 'Generate Timetable' button  
**Step 3:** Click on 'Download' button to download the Excel file
""")

# --- Campus Selection ---
campus = st.selectbox(
    "Select your campus:",
    ["New Delhi", "Gurgaon"]
)

# --- Google Sheet URLs for each campus ---
campus_urls = {
    "New Delhi": "https://docs.google.com/spreadsheets/d/1hxMVAdZM-aaHY1IDy7Hg8wLPdevSEhVx/edit?usp=sharing&ouid=106900160560444308561&rtpof=true&sd=true",
    "Gurgaon": "https://docs.google.com/spreadsheets/d/1owRJJCGwo9J5o24grEM3IWHECb4oE2NL/edit?usp=sharing&ouid=106900160560444308561&rtpof=true&sd=true"
}

url = campus_urls[campus]

# -------------------------------------------
# File download and preprocessing (same logic)
# -------------------------------------------

# Extract the File Id from the URL
file_id = url.split('/')[-2]

# Construct the correct Download URL for xlsx format
download_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"

response = requests.get(download_url)
wb = load_workbook(BytesIO(response.content))
ws = wb.active

# Step 2: Handle merged cells - Unmerge and fill values
merged_info = []
for merged_range in list(ws.merged_cells.ranges):
    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
    value = ws.cell(row=min_row, column=min_col).value
    merged_info.append((min_row, min_col, max_row, max_col, value))

for merged_range in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merged_range))

for min_row, min_col, max_row, max_col, value in merged_info:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = value

# Step 3: Load into DataFrame
data = list(ws.values)
tt = pd.DataFrame(data)

# Step 4: Set headers
tt.columns = tt.iloc[0]
tt = tt[1:].reset_index(drop=True)

# Step 6: Remove columns with same name and same values
def drop_duplicate_named_columns(df):
    keep = []
    seen = set()
    for col in df.columns:
        col_data = tuple(df[col])
        col_key = (col, col_data)
        if col_key not in seen:
            seen.add(col_key)
            keep.append(col)
    return df[keep]

tt = drop_duplicate_named_columns(tt)
tt = tt.loc[:, ~tt.T.duplicated(keep='first')]

# Step 7: Rename first column to 'Day/time' if needed
if tt.columns[0] != 'Day/time':
    tt.rename(columns={tt.columns[0]: 'Day/time'}, inplace=True)

# Step 8: Remove duplicate (column name, row header, value)
row_header_col_index = 0
row_headers = tt.iloc[:, row_header_col_index]
seen = set()

for col_index in range(1, tt.shape[1]):
    for row_index in range(tt.shape[0]):
        row_header_value = row_headers.iat[row_index]
        cell_value = tt.iat[row_index, col_index]
        col_name = tt.columns[col_index]
        key = (col_name, row_header_value, cell_value)
        if key in seen:
            tt.iat[row_index, col_index] = np.nan
        else:
            seen.add(key)

tt.replace(to_replace=["None", None], value=np.nan, inplace=True)
tt.dropna(axis=1, how='all', inplace=True)
tt.columns = tt.columns.astype(str)

# Rename duplicate column names
new_cols = []
counter = 1
seen = set()
for col in tt.columns:
    if col in seen:
        new_cols.append(f"Unnamed {counter}")
        counter += 1
    else:
        seen.add(col)
        new_cols.append(col)
tt.columns = new_cols

unnamed_cols = [col for col in tt.columns if col.startswith('Unnamed')]
for col in unnamed_cols:
    left_col_index = tt.columns.get_loc(col) - 1
    left_col = tt.columns[left_col_index]
    for index, row in tt.iterrows():
        if pd.notna(row[col]):
            if pd.notna(row[left_col]):
                tt.loc[index, left_col] = f"{row[left_col]} {row[col]}"
            else:
                tt.loc[index, left_col] = row[col]
tt = tt.drop(columns=unnamed_cols)

# -------------------------------------------
# Dynamic subject extraction
# -------------------------------------------

# Extract all unique subjects from the first term of each timetable cell
subject_set = set()
for col in tt.columns:
    if col != "Day/time":
        for val in tt[col].dropna():
            parts = str(val).split()
            if len(parts) > 0:
                subject_set.add(parts[0])

all_subjects = sorted(subject_set)

# -------------------------------------------
# Subject selection in Streamlit
# -------------------------------------------

my_subjects = st.multiselect("Select your subjects:", all_subjects)

if st.button("Generate Timetable"):
    if not my_subjects:
        st.warning("Please select at least one subject to generate your timetable.")
    else:
        # Create an empty DataFrame for personal timetable
        personal_tt = pd.DataFrame(columns=tt.columns)

        # Populate personal timetable
        for index, row in tt.iterrows():
            for col in tt.columns:
                if col != 'Day/time':
                    cell_value = row[col]
                    if pd.notna(cell_value):
                        cell_subjects = str(cell_value).split()
                        for subject in my_subjects:
                            if subject in cell_subjects:
                                personal_tt.loc[index, col] = cell_value
                                personal_tt.loc[index, 'Day/time'] = row['Day/time']
                                break

        unique_days = tt[['Day/time']].drop_duplicates()
        personal_tt = pd.merge(unique_days, personal_tt, on='Day/time', how='left')
        personal_tt = personal_tt.groupby('Day/time').agg(lambda x: ' '.join(x.dropna())).reset_index()

        original_day_order = tt['Day/time'].drop_duplicates()
        personal_tt['Day/time'] = pd.Categorical(personal_tt['Day/time'], categories=original_day_order, ordered=True)
        personal_tt = personal_tt.sort_values('Day/time').reset_index(drop=True)

        st.dataframe(personal_tt)

        # Save to Excel (in memory)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            personal_tt.to_excel(writer, sheet_name='Personal Timetable', index=False)

            workbook = writer.book
            worksheet = writer.sheets['Personal Timetable']

            cell_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
            bold_format = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})

            for row_num in range(len(personal_tt) + 1):
                for col_num in range(len(personal_tt.columns)):
                    cell_value = personal_tt.iloc[row_num - 1, col_num] if row_num > 0 else personal_tt.columns[col_num]
                    if row_num == 0 or col_num == 0:
                        worksheet.write(row_num, col_num, cell_value, bold_format)
                    else:
                        worksheet.write(row_num, col_num, cell_value, cell_format)

            for row_num in range(len(personal_tt) + 1):
                worksheet.set_row(row_num, 32.4)
            for col_num in range(len(personal_tt.columns)):
                worksheet.set_column(col_num, col_num, 20)

        st.download_button(
            label="Download Timetable as Excel",
            data=output.getvalue(),
            file_name=f"personal_timetable_{campus.lower().replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
